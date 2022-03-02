import csv
import inspect
import logging

import softest as softest
from openpyxl import load_workbook


class Utils(softest.TestCase):
    def list_item_text(self, list1, value):
        for stop in list1:
            print(stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("assert pass")
            else:
                print("aasert fail")
        self.assert_all()

# creating the custom logger

    def custom_logger(logLevel = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("automation.log", 'w')
        formatter = logging.Formatter("%(asctime)s - %(levelname)s : %(message)s", datefmt="%d/%m/%Y %I:%M:%S %p")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        col_count = sh.max_column

        for i in range(2, row_count + 1):
            row = []
            for j in range(1, col_count + 1):
                value = sh.cell(row=i, column=j).value
                # if j == 3:
                #     datevalue = value.strftime('%d/%m/%y')   # to get the date in dd/mm/yyyy format while using excel
                #     row.append(datevalue)
                #     continue

                row.append(value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist = []
        csvdata = open(filename, 'r')
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist




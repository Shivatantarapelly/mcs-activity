from openpyxl import load_workbook


def read_data_from_excel(file_name, sheet):
    datalist = []
    wb = load_workbook(filename=file_name)
    sh = wb[sheet]
    row_count = sh.max_row
    col_count = sh.max_column

    for i in range(2, row_count + 1):
        row = []
        for j in range(1, col_count + 1):
            value = sh.cell(row=i, column=j).value
            if j == 3:
                datevalue = value.strftime("%d/%m/%y")
                row.append(datevalue)
                continue
            row.append(sh.cell(row=i, column=j).value)

        datalist.append(row)
    return datalist


n = read_data_from_excel(r"D:\Shiva\files\flask git project\selenium\TestFrameworkDemo1\testdata\exceldata.xlsx",
                         "Sheet1")
for i in n:
    print(i)

print("==========================================")

lst1 = ["()", ")(()))", "(", "(())((()())())", ")test", "hi())(", "hello))", "hello", "(hell(0))"]
# lst1 = ["hello"]
for item in lst1:
    if "(" in item or ")" in item:
        seq = False
    else:
        seq = True
    x = []
    for l in item:
        x.append(l)
    # print(x)
    if x[-1] == "(":
        seq = False
    elif "(" in x:
        for i in x:
            if i == ")":
                seq = False
                break
            elif i == "(":
                x.remove(i)
                for j in x:
                    if j == ")":
                        x.remove(j)
                        seq = True
            elif i != "(":
                x.remove(i)

    if seq:
        print(f'{item} ==> {True}')
    else:
        print(f'{item} ==> {False}')

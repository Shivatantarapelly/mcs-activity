from faker import Faker
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# fakedata = Faker('hi_IN')  # hi_IN stores the data in hindi language, en_US in english
fakedata = Faker()    # we can give the language or by default we get english data
# inserting fake data into excel sheet of 10rows and 3 columns
for i in range(1,11):
    for j in range(1,4):
        ws.cell(row=i, column=1).value = fakedata.name()
        ws.cell(row=i, column=2).value = fakedata.email()
        ws.cell(row=i, column=3).value = fakedata.address()
wb.save("fakedata.xlsx")
